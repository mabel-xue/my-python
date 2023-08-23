# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup
import pandas as pd
from api import get_caibaoshuo_stocks, get_dongfang_stocks
from utils import get_soup
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Alignment
from openpyxl.styles import Font

header={
        'Cookie': '_ga=GA1.2.1733667941.1647749329; amp_d915a9=voJF7siUVkMVPRAlsjfT4G.M2h6VHJVbFVEeVRYYWtSZXowQmJGSG9TZlp3Mg==..1ggkkb752.1ggkkbsbd.0.8.8; amp_adc4c4=SQowhrRwuW-_VIUx2_JSnq.M2h6VHJVbFVEeVRYYWtSZXowQmJGSG9TZlp3Mg==..1gpnn24st.1gpnn9t98.0.s.s; Hm_lvt_09f1f334db9a482bdf5fa1cb58cc9bf5=1683084643; acw_tc=3ccdc15b16851962787953566e65bd922fb5bc7575e1531c37b0c9b944e6ed; acw_sc__v2=64720e1cbf8391979d2dc9d0a07d5f87a5f3b429; _caibaoshuo_session=MDdwSWcyREt4Ny94WUJDUDRxWTc5SllFNy9tWlp6RTBLdVNIRTNLbk43ZDFoUk5pWEF6MDV1OHFneFdWS3FWSFNZNEZrTmhMQUhTSGV6STd6MzZYWVEza2JkcFVUaC9YSEFUYjZXbkNGVUdaSEtJNElLUVplMVliZ0FobFRkU1BWdDNOMk5yRWFMQk8yTmsvS1ByQTA5R2FXWG1pTUJpZE5hWFZLNmFoc1VYNVQyWERoV3ZVQnJDY2szaVRCbmFPZldIZ3pwVU9mVVNRRDVEK3NKSndyTDQ3UUw2eFR5R09CdTNldmtodXdWZC93Z1ROWU9mREhOcU93a0pGU2JIVzJ0YzhHbjNxc2JEcWMxaXFXMnpWR3RPMEZQRXFGQnN4UXJPbTl6M2MvcDNaZ1lRWmxXRU1aQjNiVmxUWThtdHdiYWs1L2xJY2IvczJEcG1JUzF5UnMzQkZiSWlsV29IL0lJSk9ETXJWOHROT3RuaStoR1dmZmRmUENNcUN2d2ZPLS1BT05sTVo1N1ptdnJkRXFncFdLSG13PT0%3D--efcbd128fb5706542e53ac04f3fd65a11af81997; Hm_lpvt_09f1f334db9a482bdf5fa1cb58cc9bf5=1685196323'
    }

def wrap_cell(sheet):
  for col in range(1, sheet.max_column + 1):
    cell = sheet.cell(row=2, column=col)
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

def merge_cell(sheet):
  sheet.merge_cells(start_row=1, end_row=1, start_column=2, end_column=3)
  sheet.merge_cells(start_row=1, end_row=1, start_column=4, end_column=6)
  sheet.merge_cells(start_row=1, end_row=1, start_column=7, end_column=15)
  sheet.merge_cells(start_row=1, end_row=1, start_column=16, end_column=25)
  sheet.merge_cells(start_row=1, end_row=1, start_column=26, end_column=28)
  sheet.merge_cells(start_row=1, end_row=1, start_column=29, end_column=31)
  for col in range(2, sheet.max_column + 1):
    cell = sheet.cell(row=1, column=col)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def num_cell(sheet, range):
  # 遍历指定区域的单元格
  for row in sheet[range]:
      for cell in row:
          try:
            # 将单元格的值转换为数字类型
            cell.value = float(cell.value) if cell.value is not None else None
          except UnicodeEncodeError:
            # 处理无法编码为十进制字符串的字符
            cell.value = None

industryCode='309'
file_name="{}.xlsx".format(industryCode)
# 对比数据
# codes=get_dongfang_stocks('b:BK0451+f:!50')
codes=get_caibaoshuo_stocks(industryCode,header)
# codes=[{
#   "name": '青岛啤酒',
#   "code": '600600'
# },{
#   "name": '重庆啤酒',
#   "code": '600132'
# },{
#   "name": '珠江啤酒',
#   "code": '002461'
# },{
#   "name": '惠泉啤酒',
#   "code": '600573'
# },{
#   "name": '燕京啤酒',
#   "code": '000729'
# },{
#   "name": '兰州黄河',
#   "code": '000929'
# }]
vs_df=None
writer = pd.ExcelWriter(file_name, engine='openpyxl')

for item in codes:
  url = 'https://caibaoshuo.com/companies/{}/financials'.format(item['code'])
  # print(url)
  div = get_soup(url,header).find('div', {'id': 'alkey-yearly'})
  table = div.find('table')
  df = pd.read_html(str(table), encoding='utf-8')[0]
  # print(df)
  #
  if vs_df is None or vs_df.empty:
    vs_df = df.iloc[:, 0:2].copy()
  # vs_df[item['name'].decode('utf-8')]=df['2022']
  vs_df[item['name'].decode('utf-8')]=df['2022']

  # print(vs_df)
  # 删除 趋势 列
  del df[df.columns[2]]
  df.to_excel(writer, sheet_name=item['name'].decode('utf-8'), index=False)

vs_df = vs_df.transpose()
vs_df = vs_df.reset_index()
# vs_df.index.name = None
vs_df.to_excel(writer, sheet_name='vs', index=False, header=False)


# change sheet index
wb = writer.book
sheets = wb.sheetnames
if 'vs' in sheets:
    sheet_index = sheets.index('vs')
    sheets.insert(0, sheets.pop(sheet_index))
wb._sheets = [wb[sheet_name] for sheet_name in sheets]

writer.save()

# 打开 Excel 文件
book = openpyxl.load_workbook(file_name)

# 获取 vs sheet
vs_sheet = book['vs']
num_cell(vs_sheet, 'B3:AE{}'.format(len(codes)+2))

wrap_cell(vs_sheet)

merge_cell(vs_sheet)

# set th bold
for row in vs_sheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.font = Font(bold=True)

# 保存文件
book.save(file_name)
